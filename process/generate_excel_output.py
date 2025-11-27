import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .map_inputs import generate_output
from .parse_cohorts import parse_cohorts
from .parse_responses import parse_responses


def write_to_excel(results, base_path):
    for facilitator in results:
        # Create a new workbook for each facilitator
        workbook = Workbook()
        # Remove the default sheet
        workbook.remove(workbook.active)
        
        for week in results[facilitator]:
            # Create a new worksheet for each week
            worksheet = workbook.create_sheet(title=week[:31].replace(":", ""))  # Excel sheet names have a 31 character limit
            
            # Get all unique questions for this week
            all_questions = set()
            for user in results[facilitator][week]:
                for response in results[facilitator][week][user]["responses"]:
                    all_questions.add(response["question"])
            
            all_questions = sorted(list(all_questions))
            
            # Create headers
            worksheet.cell(row=1, column=1, value="Student Name")
            for i, question in enumerate(all_questions):
                cell = worksheet.cell(row=1, column=i + 2, value=question)
                cell.font = Font(bold=True)
            
            # Make the first column header bold as well
            worksheet.cell(row=1, column=1).font = Font(bold=True)
            
            # Fill in the data
            row = 2
            for user in sorted(results[facilitator][week].keys()):
                worksheet.cell(row=row, column=1, value=user)
                
                # Create a mapping of questions to answers for this user
                user_responses = {}
                for response in results[facilitator][week][user]["responses"]:
                    user_responses[response["question"]] = response["answer"]
                
                # Fill in answers for each question
                for i, question in enumerate(all_questions):
                    answer = user_responses.get(question, "")
                    worksheet.cell(row=row, column=i + 2, value=answer)
                
                row += 1
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        path = base_path + facilitator
        os.makedirs(path, mode=0o777, exist_ok=True)
        workbook.save(path + "/" + facilitator + ".xlsx")


if __name__ == "__main__":
    base_path = os.path.dirname(os.path.abspath(__file__))

    results_file_name = base_path + "/input/responses.csv"
    cohorts_file_name = base_path + "/input/cohorts.json"
    student_id_file_name = base_path + "/input/airtable_name_to_teachable_name.json"
    output_file_name = base_path + "/output/output.json"

    parsed_input = parse_responses(results_file_name)
    parsed_mapping = parse_cohorts(cohorts_file_name)
    parsed_student_ids = parse_cohorts(student_id_file_name)

    results = generate_output(parsed_input, parsed_mapping, parsed_student_ids)

    write_to_excel(results, base_path + "/output/")